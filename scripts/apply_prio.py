#!/usr/bin/env python3
"""
Spielt aus der App exportierte a/b/c-Einstufungen zurück in source/wortliste.xlsx.

Workflow (kein Terminal nötig):
  1. In der App: Daten -> "Aktuelle a/b/c-Liste exportieren (CSV)"
  2. Die Datei prio-update-*.csv in den Ordner updates/ legen und committen (GitHub Desktop)
  3. Die Action apply-prio.yml ruft dieses Skript, aktualisiert die xlsx und löscht die CSV
  4. Der xlsx-Commit triggert build-data.yml -> words.json wird neu gebaut

Lokal manuell:  python scripts/apply_prio.py
"""
import csv, glob, os, sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "source", "wortliste.xlsx")
UPDATES = os.path.join(ROOT, "updates")

def main():
    csvs = sorted(glob.glob(os.path.join(UPDATES, "*.csv")))
    if not csvs:
        print("Keine Updates in updates/ gefunden.")
        return 0

    wb = load_workbook(XLSX)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col_es = header.index("Spanisch") + 1
    col_pr = header.index("Priorität") + 1

    # Map Spanisch -> Zeilennr
    by_es = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_es).value
        if v is not None:
            by_es[str(v).strip()] = r

    changed = 0
    for path in csvs:
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                es = (row.get("Spanisch") or "").strip()
                pr = (row.get("Prioritaet") or row.get("Priorität") or "").strip().lower()
                if pr not in ("a", "b", "c") or es not in by_es:
                    continue
                cell = ws.cell(row=by_es[es], column=col_pr)
                if str(cell.value).strip().lower() != pr:
                    cell.value = pr
                    changed += 1
        print(f"verarbeitet: {os.path.basename(path)}")

    if changed:
        wb.save(XLSX)
        print(f"{changed} Prioritäten aktualisiert -> {XLSX}")
    else:
        print("Keine Änderung gegenüber bestehender xlsx.")

    # verarbeitete CSVs entfernen
    for path in csvs:
        os.remove(path)
    print(f"{len(csvs)} CSV(s) aus updates/ entfernt.")
    return 0

if __name__ == "__main__":
    sys.exit(main())
