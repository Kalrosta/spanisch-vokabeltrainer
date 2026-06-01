#!/usr/bin/env python3
"""
Konvertiert source/wortliste.xlsx -> data/words.json + data/examples.json
Läuft in der GitHub Action bei jedem Push. Nur openpyxl als Abhängigkeit.

words.json    : Kernkarten (id, de, es, pos, prio, theme, hint)
examples.json : Beispielsätze separat, nach id ({ "12": {es, de} }), nur wo vorhanden.

Die Priorität (a/b/c) aus der xlsx ist NUR Startwert für noch nie geübte Karten.
Sobald eine Karte bewertet wurde, leitet die App a/b/c live aus dem FSRS-Zustand ab.
"""
import json, os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "source", "wortliste.xlsx")
DATA = os.path.join(ROOT, "data")

POS_ALLOWED = {"Nomen", "Verb", "Adjektiv", "Adverb", "Phrase",
               "Redewendung", "Konnektor", "Konjunktion", "Präposition", "Interjektion"}

def s(v):
    return "" if v is None else str(v).strip()

def main():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [s(h) for h in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    def col(row, name):
        i = idx.get(name)
        return s(row[i]) if i is not None and i < len(row) else ""

    words, examples = [], {}
    seen, warnings = set(), []
    next_id = 1

    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        de, es = col(row, "Deutsch"), col(row, "Spanisch")
        if not de or not es:
            continue
        if es in seen:
            warnings.append(f"Dublette übersprungen: {es}")
            continue
        seen.add(es)

        pos = col(row, "Wortart") or "Phrase"
        if pos not in POS_ALLOWED:
            warnings.append(f"Unbekannte Wortart '{pos}' bei '{es}' -> als Phrase behandelt")
            pos = "Phrase"
        prio = (col(row, "Priorität") or "c").lower()
        if prio not in ("a", "b", "c"):
            prio = "c"

        wid = next_id; next_id += 1
        words.append({
            "id": wid,
            "de": de,
            "es": es,
            "pos": pos,
            "prio": prio,
            "theme": col(row, "Themenblock") or "Sonstiges",
            "hint": col(row, "Form/Hinweis"),
        })
        ex_es, ex_de = col(row, "Beispiel_ES"), col(row, "Beispiel_DE")
        if ex_es or ex_de:
            examples[str(wid)] = {"es": ex_es, "de": ex_de}

    os.makedirs(DATA, exist_ok=True)
    with open(os.path.join(DATA, "words.json"), "w", encoding="utf-8") as f:
        json.dump(words, f, ensure_ascii=False, separators=(",", ":"))
    with open(os.path.join(DATA, "examples.json"), "w", encoding="utf-8") as f:
        json.dump(examples, f, ensure_ascii=False, separators=(",", ":"))

    themes = sorted({w["theme"] for w in words})
    print(f"words: {len(words)} | examples: {len(examples)} | themes: {len(themes)}")
    if warnings:
        print(f"warnings: {len(warnings)}")
        for w in warnings[:10]:
            print("  " + w)

if __name__ == "__main__":
    main()
